"""
app.py — Interface Streamlit — iFind Clínica v2
Compatível com Streamlit >= 1.40 (sem injeção de CSS via st.markdown).
Tema definido em .streamlit/config.toml
"""

import streamlit as st
import tempfile
import os
import platform
import shutil
import subprocess
import sys
import json
import pandas as pd
from pathlib import Path
from datetime import datetime

# ─────────────────────────────────────────────────────────────
#  VERSÃO DO SISTEMA
# ─────────────────────────────────────────────────────────────
VERSION = "2.0.0"

# ─────────────────────────────────────────────────────────────
#  PAGE CONFIG — obrigatoriamente primeiro
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="iFind — Clínica",
    page_icon="🏥",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────────────────────
#  CSS mínimo via st.html() — compatível com Streamlit 1.40+
#  Apenas ajustes que o config.toml não cobre.
# ─────────────────────────────────────────────────────────────
# NOTA: st.html() é a forma suportada no Streamlit >= 1.40
# para injetar HTML/CSS sem depender de unsafe_allow_html no markdown.
st.html("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;500;600;700&family=JetBrains+Mono:wght@400;500&display=swap');

*, *::before, *::after { box-sizing: border-box; }

/* Esconder elementos desnecessários — compatível com Streamlit 1.40–1.55+ */
[data-testid="stToolbar"]          { display: none !important; }
[data-testid="stSidebar"]          { display: none !important; }
[data-testid="stHeader"]           { display: none !important; }
[data-testid="stDecoration"]       { display: none !important; }
[data-testid="stStatusWidget"]     { display: none !important; }
header[data-testid="stHeader"]     { display: none !important; }
section[data-testid="stSidebar"]   { display: none !important; }
div[data-testid="stToolbar"]       { display: none !important; }
footer                             { display: none !important; }
#MainMenu                          { display: none !important; }
.stAppHeader                       { display: none !important; }
.stAppToolbar                      { display: none !important; }

/* Remove o espaço vazio que a barra deixa no topo */
.main .block-container             { padding-top: 1.5rem !important; }
[data-testid="stAppViewContainer"] > section:first-child { padding-top: 0 !important; }

/* Container */
.block-container { padding: 1.5rem 2rem !important; max-width: 1400px !important; }

/* Fontes globais */
html, body, [data-testid="stAppViewContainer"] {
    font-family: 'Sora', sans-serif !important;
}

/* Scrollbar */
::-webkit-scrollbar { width: 5px; height: 5px; }
::-webkit-scrollbar-track { background: rgba(0,0,0,0.05); }
::-webkit-scrollbar-thumb { background: rgba(29,158,117,0.4); border-radius: 3px; }

/* Animações */
@keyframes fadeUp {
    from { opacity: 0; transform: translateY(12px); }
    to   { opacity: 1; transform: translateY(0); }
}
@keyframes glow {
    0%, 100% { box-shadow: 0 4px 20px rgba(29,158,117,0.25); }
    50%       { box-shadow: 0 4px 32px rgba(29,158,117,0.45); }
}
@keyframes pulse {
    0%, 100% { opacity: 1; }
    50%       { opacity: 0.5; }
}
@keyframes countUp {
    from { opacity: 0; transform: scale(0.8); }
    to   { opacity: 1; transform: scale(1); }
}

/* Tabs */
[data-testid="stTabs"] > div:first-child {
    background: rgba(0,0,0,0.04) !important;
    border: 1px solid rgba(0,0,0,0.08) !important;
    border-radius: 14px !important;
    padding: 4px !important;
    gap: 2px !important;
}
[data-testid="stTabs"] button {
    border-radius: 10px !important;
    font-family: 'Sora', sans-serif !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    transition: all .2s !important;
    padding: 8px 18px !important;
}
[data-testid="stTabs"] button[aria-selected="true"] {
    background: linear-gradient(135deg, rgba(29,158,117,0.15), rgba(13,110,86,0.10)) !important;
    color: #0d6e52 !important;
    border: 1px solid rgba(29,158,117,0.25) !important;
    font-weight: 600 !important;
}
[data-testid="stTabs"] > div:nth-child(2) { border: none !important; }

/* Inputs */
[data-testid="stTextInput"] input,
[data-testid="stNumberInput"] input {
    border-radius: 10px !important;
    font-family: 'Sora', sans-serif !important;
    font-size: 13px !important;
    transition: border-color .2s, box-shadow .2s !important;
}
[data-testid="stTextInput"] input:focus,
[data-testid="stNumberInput"] input:focus {
    border-color: rgba(29,158,117,0.5) !important;
    box-shadow: 0 0 0 3px rgba(29,158,117,0.12) !important;
}

/* Labels */
[data-testid="stTextInput"] label,
[data-testid="stNumberInput"] label,
[data-testid="stSelectbox"] label,
[data-testid="stSlider"] label,
[data-testid="stRadio"] label,
[data-testid="stCheckbox"] label,
[data-testid="stFileUploader"] label {
    font-size: 12px !important;
    font-weight: 500 !important;
    font-family: 'Sora', sans-serif !important;
}

/* Botão primário */
[data-testid="stButton"] button[kind="primary"] {
    background: linear-gradient(135deg, #1D9E75, #0d6e52) !important;
    border: none !important;
    color: #fff !important;
    animation: glow 3s ease-in-out infinite !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    letter-spacing: 0.3px !important;
    border-radius: 10px !important;
    transition: all .2s !important;
}
[data-testid="stButton"] button[kind="primary"]:hover {
    background: linear-gradient(135deg, #22b585, #0f7d5e) !important;
    transform: translateY(-2px) !important;
}

/* Botões secundários */
[data-testid="stButton"] button {
    border-radius: 10px !important;
    font-family: 'Sora', sans-serif !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    transition: all .2s !important;
}
[data-testid="stButton"] button:hover {
    transform: translateY(-1px) !important;
}

/* File uploader */
[data-testid="stFileUploader"] > div {
    border-radius: 14px !important;
    transition: all .25s !important;
}

/* Progress bar */
[data-testid="stProgress"] > div > div > div {
    background: linear-gradient(90deg, #1D9E75, #5DCAA5) !important;
    border-radius: 8px !important;
    box-shadow: 0 0 12px rgba(29,158,117,0.4) !important;
    transition: width .4s ease !important;
}

/* Expander */
[data-testid="stExpander"] {
    border-radius: 12px !important;
    overflow: hidden !important;
}

/* Métricas */
[data-testid="metric-container"] {
    border-radius: 14px !important;
    padding: 16px !important;
    animation: fadeUp .5s ease both !important;
}
[data-testid="metric-container"] [data-testid="stMetricValue"] {
    font-family: 'Sora', sans-serif !important;
    font-size: 28px !important;
    font-weight: 700 !important;
    animation: countUp .4s cubic-bezier(.34,1.56,.64,1) both !important;
}
[data-testid="metric-container"] [data-testid="stMetricLabel"] {
    font-size: 11px !important;
    text-transform: uppercase !important;
    letter-spacing: .7px !important;
}

/* Alertas */
[data-testid="stAlert"] { border-radius: 12px !important; }

/* Código */
[data-testid="stCode"] {
    border-radius: 10px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 12px !important;
}

/* Download button */
[data-testid="stDownloadButton"] button {
    border-radius: 10px !important;
    font-family: 'Sora', sans-serif !important;
}

/* Markdown */
[data-testid="stMarkdownContainer"] code {
    border-radius: 4px !important;
    padding: 1px 5px !important;
    font-family: 'JetBrains Mono', monospace !important;
    font-size: 11px !important;
}

/* Animação nos blocos */
[data-testid="stVerticalBlock"] > div { animation: fadeUp .3s ease both; }
</style>
""")


# ─────────────────────────────────────────────────────────────
#  Tesseract check
# ─────────────────────────────────────────────────────────────
def _tesseract_ok() -> bool:
    if shutil.which("tesseract"):
        return True
    if platform.system().lower() == "windows":
        for exe in [
            Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe"),
            Path(r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe"),
            Path(os.environ.get("LOCALAPPDATA","")) / "Tesseract-OCR" / "tesseract.exe",
            Path(os.environ.get("APPDATA",""))      / "Tesseract-OCR" / "tesseract.exe",
        ]:
            if exe.exists():
                os.environ["PATH"] = str(exe.parent) + os.pathsep + os.environ.get("PATH","")
                return True
    portatil = Path(__file__).parent / "tesseract_bin" / "tesseract.exe"
    if portatil.exists():
        os.environ["PATH"] = str(portatil.parent) + os.pathsep + os.environ.get("PATH","")
        return True
    cfg_tess = Path(__file__).parent / "config_tesseract.py"
    if cfg_tess.exists():
        try:
            exec(cfg_tess.read_text(encoding="utf-8"), {})
            import pytesseract as _pt; _pt.get_tesseract_version(); return True
        except Exception: pass
    return False


def _verificar_tesseract():
    if _tesseract_ok(): return
    st.warning("⚙️ **Tesseract OCR não encontrado.** Iniciando instalação automática...")
    setup_path = Path(__file__).parent / "setup_tesseract.py"
    if not setup_path.exists(): setup_path = Path(__file__).parent / "setup.py"
    if not setup_path.exists():
        st.error("❌ Arquivo de instalação não encontrado."); st.stop()
    import ctypes
    def _tem_admin():
        try: return ctypes.windll.shell32.IsUserAnAdmin() != 0
        except: return False
    if platform.system().lower() == "windows" and not _tem_admin():
        res = ctypes.windll.shell32.ShellExecuteW(None,"runas",sys.executable,f'"{setup_path.resolve()}"',None,1)
        if res <= 32:
            st.error("❌ Falha ao solicitar admin."); st.stop()
        import time
        for _ in range(36):
            time.sleep(5)
            if _tesseract_ok(): break
    else:
        r = subprocess.run([sys.executable, str(setup_path)], capture_output=True, text=True, timeout=360)
        if r.returncode != 0: st.error("❌ Instalação automática falhou."); st.stop()
    if _tesseract_ok():
        import time; time.sleep(1); st.rerun()
    else:
        st.warning("⚠️ Reinicie o terminal e execute novamente."); st.stop()


_verificar_tesseract()

# ─────────────────────────────────────────────────────────────
#  Módulos do projeto
# ─────────────────────────────────────────────────────────────
try:
    from database import (
        inicializar_banco, listar_execucoes, resultados_da_execucao,
        iniciar_execucao, finalizar_execucao, salvar_resultado,
        estatisticas_gerais, listar_usuarios, criar_usuario,
        alterar_senha, desativar_usuario,
    )
    from auth      import tela_login, usuario_atual, fazer_logout, inicializar_sessao
    from processor import processar_lista, validar_planilha, ler_planilha
    from mailer    import enviar_email, enviar_relatorio_execucao, email_configurado
except ImportError as e:
    st.error(f"❌ Erro de importação: `{e}`")
    st.info("Execute: `pip install -r requirements.txt`")
    st.stop()

# ─────────────────────────────────────────────────────────────
#  Helpers visuais
# ─────────────────────────────────────────────────────────────
def badge(texto: str, cor: str = "verde") -> str:
    paleta = {
        "verde":   ("#d1fae5", "#065f46", "#6ee7b7"),
        "vermelho":("#fee2e2", "#991b1b", "#fca5a5"),
        "azul":    ("#dbeafe", "#1e40af", "#93c5fd"),
        "amarelo": ("#fef3c7", "#92400e", "#fcd34d"),
        "cinza":   ("#f3f4f6", "#374151", "#d1d5db"),
    }
    bg, fg, borda = paleta.get(cor, paleta["cinza"])
    return (f'<span style="display:inline-flex;align-items:center;gap:4px;background:{bg};'
            f'color:{fg};border:1px solid {borda};border-radius:20px;padding:3px 10px;'
            f'font-size:11px;font-weight:600;font-family:\'Sora\',sans-serif;">{texto}</span>')


def secao_titulo(icone: str, titulo: str, subtitulo: str = "") -> None:
    sub = (f'<div style="font-size:12px;color:#6b7280;margin-top:2px;">'
           f'{subtitulo}</div>') if subtitulo else ""
    st.markdown(
        f"""<div style="display:flex;align-items:center;gap:12px;margin-bottom:1rem;
            animation:fadeUp .4s ease both;">
            <div style="width:38px;height:38px;
                background:linear-gradient(135deg,rgba(29,158,117,0.15),rgba(13,110,86,0.08));
                border:1px solid rgba(29,158,117,0.3);border-radius:10px;
                display:flex;align-items:center;justify-content:center;
                font-size:18px;flex-shrink:0;">{icone}</div>
            <div>
                <div style="font-size:16px;font-weight:600;color:#111827;">{titulo}</div>
                {sub}
            </div>
        </div>""",
        unsafe_allow_html=True,
    )


def tabela_custom(linhas_html: str, colunas: list[str]) -> None:
    ths = "".join(
        f'<th style="padding:10px 14px;text-align:left;font-size:10px;'
        f'color:#6b7280;text-transform:uppercase;letter-spacing:.6px;'
        f'font-weight:600;">{c}</th>'
        for c in colunas
    )
    st.markdown(
        f"""<div style="background:#fff;border:1px solid #e5e7eb;
            border-radius:14px;overflow:hidden;margin-top:.5rem;overflow-x:auto;
            box-shadow:0 1px 4px rgba(0,0,0,0.06);">
            <table style="width:100%;border-collapse:collapse;font-family:'Sora',sans-serif;">
                <thead><tr style="background:#f9fafb;
                    border-bottom:1px solid #e5e7eb;">{ths}</tr></thead>
                <tbody>{linhas_html}</tbody>
            </table>
        </div>""",
        unsafe_allow_html=True,
    )


# ─────────────────────────────────────────────────────────────
#  App principal
# ─────────────────────────────────────────────────────────────
# ─────────────────────────────────────────────────────────────
#  AUTO-UPDATE
# ─────────────────────────────────────────────────────────────
def _verificar_update():
    """
    Verifica silenciosamente se há update disponível.
    Exibe notificação na interface se houver — não interrompe o uso.
    O update só é aplicado se o usuário confirmar (exceto se obrigatorio=True).
    """
    # Só verifica uma vez por sessão
    if st.session_state.get("_update_verificado"):
        return
    st.session_state["_update_verificado"] = True

    try:
        from updater import verificar_versao_disponivel, executar_update, VERSION as UPDATER_VERSION
    except ImportError:
        return  # updater.py não encontrado — ignora silenciosamente

    try:
        dados = verificar_versao_disponivel(timeout=5)
    except Exception:
        return  # sem internet ou servidor fora — ignora

    if not dados:
        return  # já atualizado

    versao_nova = dados.get("version", "?")
    notas       = dados.get("notas", "")
    obrigatorio = dados.get("obrigatorio", False)

    if obrigatorio:
        # Update obrigatório: aplica imediatamente com aviso
        with st.spinner(f"🔄 Aplicando atualização obrigatória v{versao_nova}..."):
            sucesso = executar_update(dados, silencioso=True)
        if sucesso:
            st.success(f"✅ Sistema atualizado para v{versao_nova}! Recarregando...")
            import time; time.sleep(2)
            st.rerun()
        else:
            st.warning("⚠️ Falha na atualização automática. Continue usando normalmente.")
        return

    # Update opcional: mostra notificação não intrusiva
    if not st.session_state.get("_update_aceito") and not st.session_state.get("_update_recusado"):
        nota_html = f"<br><small style='color:#6b7280'>{notas}</small>" if notas else ""
        st.info(
            f"🆕 **Nova versão disponível: v{versao_nova}**{nota_html}  \n"
            f"Versão atual: v{VERSION}",
            icon="🔄"
        )
        col_sim, col_nao, col_esp = st.columns([1, 1, 4])
        with col_sim:
            if st.button("⬆️ Atualizar agora", key="_btn_update_sim", type="primary"):
                st.session_state["_update_aceito"] = True
                st.session_state["_dados_update"]  = dados
                st.rerun()
        with col_nao:
            if st.button("Agora não", key="_btn_update_nao"):
                st.session_state["_update_recusado"] = True
                st.rerun()

    # Executa o update se o usuário confirmou
    if st.session_state.get("_update_aceito"):
        dados_up = st.session_state.get("_dados_update", dados)
        with st.spinner(f"🔄 Baixando e aplicando v{versao_nova}..."):
            sucesso = executar_update(dados_up, silencioso=True)
        if sucesso:
            st.success(f"✅ Atualizado para v{versao_nova}! A página vai recarregar...")
            st.session_state.pop("_update_aceito", None)
            st.session_state.pop("_dados_update", None)
            import time; time.sleep(2)
            st.rerun()
        else:
            st.error("❌ Falha na atualização. Tente novamente mais tarde.")
            st.session_state.pop("_update_aceito", None)


def _main():
    # ── Verificação de update (rápida, não bloqueia se sem internet) ──
    _verificar_update()

    inicializar_banco()
    tela_login()

    usuario = usuario_atual()
    if usuario is None:
        return

    CONFIG_PATH = Path(__file__).parent / "config.json"

    def carregar_config() -> dict:
        defaults = {
            "smtp": {}, "drive_raiz": "", "pasta_destino": "",
            "threshold_fuzzy": 80, "filtrar_aso": False,
            "enviar_email_auto": False, "email_relatorio": "",
            "modo_extracao": "auto", "dpi_ocr": 150, "max_workers": 4, "varredura_total": False,
        }
        if not CONFIG_PATH.exists():
            return defaults
        try:
            d = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
            defaults.update(d)
            return defaults
        except Exception:
            return defaults

    def salvar_config(cfg: dict):
        CONFIG_PATH.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")

    cfg = carregar_config()

    if "input_drive_raiz" not in st.session_state:
        st.session_state["input_drive_raiz"] = cfg.get("drive_raiz", "")
    if "input_pasta_destino" not in st.session_state:
        st.session_state["input_pasta_destino"] = cfg.get("pasta_destino", "")

    # ── HEADER ──────────────────────────────────────────────
    nome_usuario = usuario.get("nome", "Usuário")
    iniciais     = "".join(p[0].upper() for p in nome_usuario.split()[:2]) or "U"

    col_header, col_sair = st.columns([11, 1])
    with col_header:
        st.markdown(
            f"""<div style="display:flex;align-items:center;justify-content:space-between;
                padding:12px 20px;
                background:#fff;
                border:1px solid #e5e7eb;border-radius:16px;
                margin-bottom:1rem;animation:fadeUp .4s ease both;
                box-shadow:0 1px 4px rgba(0,0,0,0.06);">
                <div style="display:flex;align-items:center;gap:12px;">
                    <div style="width:40px;height:40px;
                        background:linear-gradient(135deg,#1D9E75,#0d6e52);
                        border-radius:11px;display:flex;align-items:center;
                        justify-content:center;font-size:20px;
                        box-shadow:0 4px 16px rgba(29,158,117,0.25);">🏥</div>
                    <div>
                        <div style="font-size:16px;font-weight:700;color:#111827;letter-spacing:-.3px;">
                            iFind Clínica</div>
                        <div style="font-size:11px;color:#9ca3af;margin-top:1px;">
                            Busca automática em documentos PDF</div>
                    </div>
                </div>
                <div style="display:flex;align-items:center;gap:8px;
                    background:#f9fafb;
                    border:1px solid #e5e7eb;
                    border-radius:20px;padding:6px 14px 6px 8px;">
                    <div style="width:28px;height:28px;
                        background:linear-gradient(135deg,#1D9E75,#5DCAA5);
                        border-radius:50%;display:flex;align-items:center;
                        justify-content:center;font-size:11px;font-weight:700;color:#fff;">
                        {iniciais}</div>
                    <span style="font-size:13px;color:#374151;
                        font-weight:500;">{nome_usuario}</span>
                </div>
            </div>""",
            unsafe_allow_html=True,
        )
    with col_sair:
        st.markdown("<div style='margin-top:12px'></div>", unsafe_allow_html=True)
        if st.button("Sair →", key="btn_logout"):
            fazer_logout(); st.rerun()

    # ── TABS ────────────────────────────────────────────────
    _uid_hist = None if usuario.get("admin") else usuario.get("id")
    _n_exec   = len(listar_execucoes(usuario_id=_uid_hist, limite=999))
    _badge_hist = f"  ·  {_n_exec}" if _n_exec else ""

    aba_busca, aba_historico, aba_stats, aba_config = st.tabs([
        "🔍  Busca",
        f"📋  Histórico{_badge_hist}",
        "📊  Estatísticas",
        "⚙️  Configurações",
    ])

    # ════════════════════════════════════════════════════════
    #  ABA 1 — BUSCA
    # ════════════════════════════════════════════════════════
    with aba_busca:
        secao_titulo("🔍", "Nova busca", "Configure os parâmetros e inicie o processamento")

        def _cb_drive():
            from processor import abrir_seletor_pasta
            p = abrir_seletor_pasta("Pasta raiz do drive")
            if p: st.session_state["input_drive_raiz"] = p

        def _cb_destino():
            from processor import abrir_seletor_pasta
            p = abrir_seletor_pasta("Pasta de destino")
            if p: st.session_state["input_pasta_destino"] = p

        col1, col2 = st.columns(2, gap="medium")

        # ── Coluna esquerda ──────────────────────────────────
        with col1:
            st.markdown("##### 📄 Planilha de pacientes")
            arquivo_excel = st.file_uploader(
                "Arraste ou clique para enviar",
                type=["xlsx", "xls"],
                help="Colunas detectadas automaticamente.",
                label_visibility="collapsed",
            )

            if arquivo_excel:
                try:
                    import tempfile as _tf
                    with _tf.NamedTemporaryFile(delete=False, suffix=".xlsx") as _t:
                        _t.write(arquivo_excel.read()); _tp = _t.name
                    arquivo_excel.seek(0)
                    _regs   = ler_planilha(_tp); os.unlink(_tp)
                    _avisos = validar_planilha(_regs)

                    # Banner de sucesso via componente nativo (sem HTML manual)
                    st.success(f"✓ **{len(_regs)} registros detectados**")

                    # Avisos só aparecem se houver
                    for _av in _avisos[:3]:
                        st.warning(f"⚠ {_av['msg']}")

                    # Preview completo com st.dataframe nativo
                    with st.expander("👁️ Prévia da planilha", expanded=True):
                        _prev_df = pd.DataFrame([
                            {
                                "Nome":   r.get("nome", ""),
                                "Data":   str(r.get("data", "")),
                                "E-mail": r.get("email", "") or "—",
                            }
                            for r in _regs
                        ])
                        st.dataframe(
                            _prev_df,
                            use_container_width=True,
                            hide_index=True,
                            height=min(36 * len(_prev_df) + 38, 320),
                        )
                        st.caption(f"Exibindo {len(_regs)} registros")

                except Exception as _ex:
                    st.error(f"❌ Erro ao ler planilha: {_ex}")

            st.markdown("<div style='margin-top:14px'></div>", unsafe_allow_html=True)
            st.markdown("##### 📁 Caminhos")

            _c1, _c2 = st.columns([6, 1])
            with _c1:
                drive_raiz = st.text_input(
                    "Pasta raiz do drive",
                    placeholder="Ex: Z:\\Procedimentos",
                    key="input_drive_raiz",
                )
            with _c2:
                st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                st.button("🗂️", key="btn_drive", help="Selecionar pelo explorador", on_click=_cb_drive)

            _c3, _c4 = st.columns([6, 1])
            with _c3:
                pasta_destino = st.text_input(
                    "Pasta de destino",
                    placeholder="Ex: C:\\Resultados",
                    key="input_pasta_destino",
                )
            with _c4:
                st.markdown("<div style='margin-top:28px'></div>", unsafe_allow_html=True)
                st.button("🗂️", key="btn_destino", help="Selecionar pelo explorador", on_click=_cb_destino)

        # ── Coluna direita ───────────────────────────────────
        with col2:
            st.markdown("##### ⚙️ Configurações da busca")

            threshold = st.slider(
                "🎯 Sensibilidade fuzzy",
                min_value=60, max_value=100,
                value=cfg.get("threshold_fuzzy", 80), step=5,
                help="100 = exato  |  80 = tolera erros de OCR  |  60 = permissivo",
            )

            _MODOS = {
                "auto":   "🔄 Automático — detecta tipo do PDF",
                "nativo": "⚡ Só texto nativo — PDFs digitais (rápido)",
                "ocr":    "🔍 Só OCR — PDFs escaneados (lento)",
            }
            modo_extracao = st.radio(
                "📄 Modo de extração",
                options=list(_MODOS.keys()),
                format_func=lambda k: _MODOS[k],
                index=list(_MODOS.keys()).index(cfg.get("modo_extracao", "auto")),
            )

            _cDPI, _cWRK = st.columns(2)
            with _cDPI:
                dpi_ocr = st.select_slider(
                    "🖼️ Qualidade OCR",
                    options=[150, 200, 300],
                    value=cfg.get("dpi_ocr", 150),
                    format_func=lambda v: {150:"⚡ 150 DPI", 200:"⚖️ 200 DPI", 300:"🔬 300 DPI"}[v],
                    disabled=(modo_extracao == "nativo"),
                )
            with _cWRK:
                max_workers = st.slider(
                    "🧵 Threads",
                    min_value=1, max_value=8,
                    value=cfg.get("max_workers", 4),
                    help="PDFs processados em paralelo. Mais threads = mais rápido.",
                )

            st.markdown("<div style='margin-top:4px'></div>", unsafe_allow_html=True)
            _cASO, _cMAIL = st.columns(2)
            with _cASO:
                filtrar_aso = st.checkbox(
                    "🏥 Apenas ASO",
                    value=cfg.get("filtrar_aso", False),
                    help="Ignora páginas que não sejam Atestado de Saúde Ocupacional.",
                )
            with _cMAIL:
                enviar_email_auto = st.checkbox(
                    "📧 E-mail automático",
                    value=cfg.get("enviar_email_auto", False),
                )

            varredura_total = st.checkbox(
                "🔎 Varredura total (buscar em todos os PDFs)",
                value=cfg.get("varredura_total", False),
                help=(
                    "Ignora as datas da planilha e varre TODOS os PDFs recursivamente "
                    "a partir da pasta raiz.\n\n"
                    "Use quando a planilha tem só nomes, ou quando a estrutura de pastas "
                    "não segue o padrão mês/dia.\n\n"
                    "⚠️ Processo lento — DPI é forçado para 100 automaticamente."
                ),
            )
            if varredura_total:
                st.warning(
                    "⚠️ **Varredura total ativada** — todos os PDFs serão processados. "
                    "Pode levar vários minutos dependendo do volume de arquivos."
                )

            email_relatorio = ""
            if enviar_email_auto:
                email_relatorio = st.text_input(
                    "E-mail para relatório",
                    value=cfg.get("email_relatorio", ""),
                    placeholder="gestor@clinica.com.br",
                )

        st.markdown("<div style='margin-top:10px'></div>", unsafe_allow_html=True)
        iniciar = st.button("▶  Iniciar busca", type="primary", use_container_width=True)

        # ── Processamento ────────────────────────────────────
        if iniciar:
            erros = []
            if not arquivo_excel:         erros.append("Faça upload da planilha Excel.")
            if not drive_raiz.strip():    erros.append("Informe o caminho do drive.")
            if not pasta_destino.strip(): erros.append("Informe a pasta de destino.")
            if erros:
                for e in erros: st.error(f"❌ {e}")
                st.stop()

            cfg.update({
                "drive_raiz": drive_raiz.strip(), "pasta_destino": pasta_destino.strip(),
                "threshold_fuzzy": threshold, "filtrar_aso": filtrar_aso,
                "enviar_email_auto": enviar_email_auto, "email_relatorio": email_relatorio,
                "modo_extracao": modo_extracao, "dpi_ocr": dpi_ocr, "max_workers": max_workers, "varredura_total": varredura_total,
            })
            salvar_config(cfg)

            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
                tmp.write(arquivo_excel.read())
                caminho_tmp = tmp.name

            execucao_id = iniciar_execucao(usuario.get("id"), drive_raiz.strip(), pasta_destino.strip())

            st.markdown(
                """<div style="font-size:14px;font-weight:600;color:#111827;
                    margin:1rem 0 .8rem;display:flex;align-items:center;gap:8px;">
                    <span style="animation:pulse 1.5s ease infinite;display:inline-block;">⚡</span>
                    Processando...
                </div>""",
                unsafe_allow_html=True,
            )

            barra         = st.progress(0, text="Iniciando...")
            slot_status   = st.empty()
            slot_metricas = st.empty()
            slot_parar    = st.empty()
            slot_log      = st.empty()

            import time as _time
            _t0    = _time.time()
            _state = {"ok": 0, "erro": 0, "aviso": 0, "log": [], "last_log": "", "pct_last": -1, "cb_count": 0}
            st.session_state["parar_busca"] = False

            _COR = {
                "ok":        ("#065f46", "#059669", "#ecfdf5"),
                "erro":      ("#991b1b", "#dc2626", "#fef2f2"),
                "aviso":     ("#92400e", "#d97706", "#fffbeb"),
                "info":      ("#1e40af", "#3b82f6", "#eff6ff"),
                "pdf_salvo": ("#065f46", "#059669", "#ecfdf5"),
            }
            _ICONE = {"ok":"✅","erro":"❌","aviso":"⚠️","info":"🔍","pdf_salvo":"📄"}
            _ETAPAS_SYS = {"Planilha lida","Concluído","Pasta localizada",
                           "Agrupamento concluído","Retomando execução"}

            def cb(prog: float, etapa: str, detalhe: str = "", status: str = "info"):
                if st.session_state.get("parar_busca"):
                    return
                pct     = int(prog * 100)
                elapsed = _time.time() - _t0
                barra.progress(min(prog, 1.0), text=f"{pct}%")

                fg, accent, bg = _COR.get(status, _COR["info"])
                icone = _ICONE.get(status, "🔍")
                det_h = (f'<div style="font-size:11px;color:{accent};margin-top:3px;">'
                         f'{detalhe}</div>') if detalhe else ""
                slot_status.markdown(
                    f'<div style="background:{bg};border-left:3px solid {accent};'
                    f'border-radius:0 10px 10px 0;padding:10px 16px;margin:4px 0;">'
                    f'<span style="font-weight:600;color:{fg};">{icone} {etapa}</span>{det_h}</div>',
                    unsafe_allow_html=True,
                )

                if status in ("ok","pdf_salvo") and etapa not in _ETAPAS_SYS:
                    _state["ok"] += 1
                elif status == "erro":   _state["erro"] += 1
                elif status == "aviso":  _state["aviso"] += 1

                mins, secs = divmod(int(elapsed), 60)
                tempo_str = f"{mins}m {secs:02d}s" if mins else f"{secs}s"
                slot_metricas.markdown(
                    f'<div style="display:flex;gap:20px;padding:8px 0;font-family:\'Sora\',sans-serif;">'
                    f'<span style="font-size:13px;"><span style="color:#059669;font-weight:700;">{_state["ok"]}</span>'
                    f'<span style="color:#6b7280;font-size:11px;margin-left:3px;">encontrados</span></span>'
                    f'<span style="font-size:13px;"><span style="color:#dc2626;font-weight:700;">{_state["erro"]}</span>'
                    f'<span style="color:#6b7280;font-size:11px;margin-left:3px;">erros</span></span>'
                    f'<span style="font-size:13px;"><span style="color:#d97706;font-weight:700;">{_state["aviso"]}</span>'
                    f'<span style="color:#6b7280;font-size:11px;margin-left:3px;">não localizados</span></span>'
                    f'<span style="font-size:13px;color:#374151;font-weight:600;">⏱ {tempo_str}</span>'
                    f'</div>',
                    unsafe_allow_html=True,
                )

                with slot_parar:
                    _state["cb_count"] += 1
                    _btn_key = f"parar_{_state['cb_count']}"
                    if st.button("⏹  Parar busca", key=_btn_key, use_container_width=False):
                        st.session_state["parar_busca"] = True

                linha = f"{icone} {etapa}" + (f"  —  {detalhe}" if detalhe else "")
                if linha != _state["last_log"]:
                    _state["log"].append(linha)
                    _state["last_log"] = linha
                slot_log.code("\n".join(_state["log"][-25:]), language=None)

            try:
                resultados = processar_lista(
                    caminho_excel   = caminho_tmp,
                    drive_raiz      = drive_raiz.strip(),
                    pasta_destino   = pasta_destino.strip(),
                    threshold_fuzzy = threshold,
                    callback        = cb,
                    modo_extracao   = modo_extracao,
                    dpi_ocr         = dpi_ocr,
                    filtrar_aso     = filtrar_aso,
                    max_workers     = max_workers,
                    varredura_total = varredura_total,
                )
            except Exception as e:
                st.error(f"❌ Erro crítico: {e}")
                os.unlink(caminho_tmp)
                st.stop()

            os.unlink(caminho_tmp)

            total       = len(resultados)
            encontrados = sum(1 for r in resultados if r["encontrado"])
            for r in resultados:
                salvar_resultado(execucao_id, r)
            finalizar_execucao(execucao_id, total, encontrados)

            barra.progress(1.0, text="100% — Concluído!")
            slot_status.empty()
            slot_parar.empty()

            # Notificação sonora
            st.markdown(
                """<script>
                try {
                    const a = new AudioContext(), o = a.createOscillator(), g = a.createGain();
                    o.connect(g); g.connect(a.destination);
                    o.frequency.setValueAtTime(880, a.currentTime);
                    o.frequency.setValueAtTime(1100, a.currentTime + 0.12);
                    g.gain.setValueAtTime(0.12, a.currentTime);
                    g.gain.exponentialRampToValueAtTime(0.001, a.currentTime + 0.45);
                    o.start(); o.stop(a.currentTime + 0.45);
                } catch(e) {}
                </script>""",
                unsafe_allow_html=True,
            )

            # E-mails automáticos
            if enviar_email_auto and email_configurado():
                with st.spinner("Enviando e-mails..."):
                    enviados = 0
                    for r in resultados:
                        if r["encontrado"] and r.get("email"):
                            ok_m, _ = enviar_email(
                                destinatario=r["email"],
                                assunto=f"Documento encontrado — {r['nome']}",
                                corpo=(f"<p>Prezado(a), o documento de <strong>{r['nome']}</strong> "
                                       f"referente a <strong>{r['data']}</strong> foi localizado.</p>"),
                                caminho_pdf=r["arquivo"],
                            )
                            if ok_m: enviados += 1
                    if email_relatorio:
                        enviar_relatorio_execucao(
                            email_relatorio, total, encontrados,
                            total-encontrados, usuario.get("nome",""),
                        )
                st.success(f"📧 {enviados} e-mail(s) enviado(s).")

            # ── Resultados ───────────────────────────────────
            st.markdown("<div style='margin-top:1.5rem'></div>", unsafe_allow_html=True)
            secao_titulo("📊", "Resultados", "Resumo do processamento")

            nao_enc = total - encontrados
            taxa    = round(encontrados / total * 100) if total > 0 else 0

            _mc1, _mc2, _mc3, _mc4 = st.columns(4)
            _mc1.metric("Total processados",   total)
            _mc2.metric("✅ Encontrados",       encontrados)
            _mc3.metric("❌ Não encontrados",   nao_enc)
            _mc4.metric("🎯 Taxa de sucesso",   f"{taxa}%")

            st.markdown("<div style='margin-top:.8rem'></div>", unsafe_allow_html=True)

            # Botões de ação
            _ba1, _ba2, _ba3 = st.columns(3)
            with _ba1:
                if st.button("📂  Abrir pasta de destino", use_container_width=True):
                    try:
                        _p = Path(pasta_destino.strip())
                        if platform.system() == "Windows": os.startfile(str(_p))
                        elif platform.system() == "Darwin": subprocess.run(["open", str(_p)])
                        else: subprocess.run(["xdg-open", str(_p)])
                    except Exception as _e:
                        st.error(f"Não foi possível abrir: {_e}")

            df = pd.DataFrame(resultados).rename(columns={
                "nome":"Nome","data":"Data","email":"E-mail","encontrado":"Encontrado",
                "arquivo":"Arquivo gerado","erro":"Observação","score_fuzzy":"Score",
            })

            with _ba2:
                try:
                    import io as _io
                    import openpyxl as _xl
                    from openpyxl.styles import PatternFill, Font, Alignment
                    _wb = _xl.Workbook(); _ws = _wb.active; _ws.title = "Resultados"
                    _cabs = ["Nome","Data","E-mail","Encontrado","Score","Arquivo","Observação"]
                    for _ci, _cab in enumerate(_cabs, 1):
                        _cell = _ws.cell(row=1, column=_ci, value=_cab)
                        _cell.fill = PatternFill("solid", fgColor="0D1F2D")
                        _cell.font = Font(color="1D9E75", bold=True, size=11)
                        _cell.alignment = Alignment(horizontal="center")
                    _fv  = PatternFill("solid", fgColor="E1F5EE")
                    _fv2 = PatternFill("solid", fgColor="FCEBEB")
                    for _ri, _r in enumerate(resultados, 2):
                        _vals = [_r.get("nome",""), _r.get("data",""), _r.get("email",""),
                                 "Sim" if _r.get("encontrado") else "Não",
                                 f"{_r.get('score_fuzzy',0):.0f}%",
                                 _r.get("arquivo",""), _r.get("erro","")]
                        _fi = _fv if _r.get("encontrado") else _fv2
                        for _ci, _v in enumerate(_vals, 1):
                            _c2 = _ws.cell(row=_ri, column=_ci, value=_v)
                            _c2.fill = _fi
                    for _col in _ws.columns:
                        _ws.column_dimensions[_col[0].column_letter].width = min(
                            max(len(str(_c.value or "")) for _c in _col)+4, 50)
                    _buf = _io.BytesIO(); _wb.save(_buf)
                    st.download_button(
                        "⬇️  Baixar Excel (.xlsx)",
                        data=_buf.getvalue(),
                        file_name=f"ifind_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                except Exception:
                    st.info("openpyxl não disponível para Excel.")

            with _ba3:
                _csv = df.to_csv(index=False).encode("utf-8-sig")
                st.download_button(
                    "⬇️  Baixar CSV",
                    _csv,
                    f"ifind_{datetime.now().strftime('%Y%m%d_%H%M')}.csv",
                    "text/csv",
                    use_container_width=True,
                )

            # Tabela de resultados customizada
            st.markdown("<div style='margin-top:.8rem'></div>", unsafe_allow_html=True)
            _linhas = ""
            for _r in resultados:
                _enc  = _r.get("encontrado", False)
                _sc   = _r.get("score_fuzzy", 0)
                _obs  = _r.get("erro", "")
                if _enc:
                    _st_h  = badge("✓ Encontrado", "verde")
                    _sc_h  = f'<span style="color:#059669;font-weight:700;font-size:13px;">{_sc:.0f}%</span>'
                elif "ASO" in _obs:
                    _st_h  = badge("⊘ Não é ASO", "amarelo")
                    _sc_h  = '<span style="color:#9ca3af;">—</span>'
                else:
                    _st_h  = badge("✗ Não encontrado", "vermelho")
                    _sc_h  = '<span style="color:#9ca3af;">—</span>'
                _obs_t = (_obs[:55]+"…") if len(_obs) > 55 else _obs
                _linhas += (
                    f'<tr style="border-bottom:1px solid #f3f4f6;">'
                    f'<td style="padding:10px 14px;color:#111827;font-weight:500;font-size:13px;">{_r.get("nome","")}</td>'
                    f'<td style="padding:10px 14px;color:#6b7280;font-size:12px;">{_r.get("data","")}</td>'
                    f'<td style="padding:10px 14px;">{_st_h}</td>'
                    f'<td style="padding:10px 14px;">{_sc_h}</td>'
                    f'<td style="padding:10px 14px;color:#9ca3af;font-size:11px;'
                    f'font-family:\'JetBrains Mono\',monospace;">{_obs_t}</td>'
                    f'</tr>'
                )
            tabela_custom(_linhas, ["Nome","Data","Status","Score","Observação"])

            if nao_enc > 0:
                st.markdown("<div style='margin-top:.8rem'></div>", unsafe_allow_html=True)
                st.warning(f"⚠️ {nao_enc} registro(s) não encontrado(s). Verifique a coluna Observação.")

    # ════════════════════════════════════════════════════════
    #  ABA 2 — HISTÓRICO
    # ════════════════════════════════════════════════════════
    with aba_historico:
        secao_titulo("📋", "Histórico de execuções", "Todas as buscas realizadas")

        _uid = None if usuario.get("admin") else usuario.get("id")
        execucoes = listar_execucoes(usuario_id=_uid, limite=100)

        if not execucoes:
            st.markdown(
                """<div style="text-align:center;padding:3rem;color:#9ca3af;">
                    <div style="font-size:40px;margin-bottom:12px;">📋</div>
                    <div style="font-size:14px;">Nenhuma execução ainda.</div>
                </div>""",
                unsafe_allow_html=True,
            )
        else:
            _hf1, _hf2, _hf3 = st.columns(3)
            with _hf1:
                _nomes_u = sorted(set(e.get("usuario_nome","") for e in execucoes))
                _fu = st.selectbox("Usuário", ["Todos"] + _nomes_u)
            with _hf2:
                _ft = st.select_slider("Taxa mínima", options=[0,25,50,75,100], value=0,
                                        format_func=lambda v: f"{v}%")
            with _hf3:
                _fo = st.selectbox("Ordenar por", ["Mais recente","Mais antigo","Maior taxa","Menor taxa"])

            _execs = [e for e in execucoes
                      if (_fu == "Todos" or e.get("usuario_nome") == _fu)
                      and (e.get("encontrados",0) / max(e.get("total",1),1) * 100) >= _ft]
            if _fo == "Mais antigo":
                _execs.sort(key=lambda e: e.get("inicio",""))
            elif _fo == "Maior taxa":
                _execs.sort(key=lambda e: e.get("encontrados",0)/max(e.get("total",1),1), reverse=True)
            elif _fo == "Menor taxa":
                _execs.sort(key=lambda e: e.get("encontrados",0)/max(e.get("total",1),1))

            st.markdown(f"<div style='font-size:12px;color:#6b7280;margin:.5rem 0;'>"
                        f"{len(_execs)} execução(ões)</div>", unsafe_allow_html=True)

            _lh = ""
            for _e in _execs[:50]:
                _tot = _e.get("total",0); _enc = _e.get("encontrados",0)
                _tx  = round(_enc/_tot*100) if _tot else 0
                _ct  = "#059669" if _tx>=80 else "#d97706" if _tx>=50 else "#dc2626"
                _lh += (
                    f'<tr style="border-bottom:1px solid #f3f4f6;">'
                    f'<td style="padding:10px 14px;color:#9ca3af;font-size:12px;'
                    f'font-family:\'JetBrains Mono\',monospace;">#{_e.get("id","")}</td>'
                    f'<td style="padding:10px 14px;color:#111827;font-size:12px;">{_e.get("usuario_nome","")}</td>'
                    f'<td style="padding:10px 14px;color:#6b7280;font-size:12px;">{str(_e.get("inicio",""))[:16]}</td>'
                    f'<td style="padding:10px 14px;color:#6b7280;font-size:12px;">{str(_e.get("fim",""))[:16] or "—"}</td>'
                    f'<td style="padding:10px 14px;text-align:center;color:#111827;font-weight:600;">{_tot}</td>'
                    f'<td style="padding:10px 14px;text-align:center;color:#059669;font-weight:600;">{_enc}</td>'
                    f'<td style="padding:10px 14px;text-align:center;color:{_ct};font-weight:700;">{_tx}%</td>'
                    f'</tr>'
                )
            tabela_custom(_lh, ["#","Usuário","Início","Fim","Total","Encontrados","Taxa"])

            st.markdown("<div style='margin-top:1.2rem'></div>", unsafe_allow_html=True)
            secao_titulo("🔎", "Detalhes de execução")
            _ids = [e["id"] for e in _execs]
            if _ids:
                _sel = st.selectbox("Execução", _ids, format_func=lambda x: f"#{x}")
                if _sel:
                    _res = resultados_da_execucao(_sel)
                    if _res:
                        _dfr = pd.DataFrame(_res).rename(columns={
                            "nome":"Nome","data":"Data","encontrado":"Encontrado",
                            "arquivo":"Arquivo","erro":"Observação","score_fuzzy":"Score",
                        })
                        def _cor_hist(row):
                            c = "rgba(29,158,117,0.08)" if row["Encontrado"] else "rgba(226,75,74,0.06)"
                            return [f"background-color:{c}"] * len(row)
                        st.dataframe(
                            _dfr[["Nome","Data","Encontrado","Score","Observação"]].style.apply(_cor_hist, axis=1),
                            use_container_width=True, hide_index=True,
                        )
                        st.download_button(
                            f"⬇️ Exportar #{_sel}",
                            _dfr.to_csv(index=False).encode("utf-8-sig"),
                            f"exec_{_sel}.csv", "text/csv",
                        )
                    else:
                        st.info("Sem detalhes para esta execução.")

    # ════════════════════════════════════════════════════════
    #  ABA 3 — ESTATÍSTICAS
    # ════════════════════════════════════════════════════════
    with aba_stats:
        secao_titulo("📊", "Estatísticas de uso", "Visão geral do sistema")

        try:
            import plotly.express as px; PLOTLY = True
        except ImportError:
            PLOTLY = False; st.warning("Instale plotly: `pip install plotly`")

        stats = estatisticas_gerais()
        _sc1,_sc2,_sc3,_sc4 = st.columns(4)
        _sc1.metric("Execuções",          stats["total_execucoes"])
        _sc2.metric("Total de buscas",    stats["total_buscas"])
        _sc3.metric("Encontrados",        stats["total_encontrados"])
        _sc4.metric("Taxa média",         f"{stats['taxa_sucesso']}%")

        _layout_light = dict(
            plot_bgcolor  = "rgba(0,0,0,0)",
            paper_bgcolor = "rgba(0,0,0,0)",
            font          = dict(color="#374151", family="Sora"),
            xaxis         = dict(gridcolor="#f3f4f6", linecolor="#e5e7eb"),
            yaxis         = dict(gridcolor="#f3f4f6", linecolor="#e5e7eb"),
            legend        = dict(bgcolor="rgba(0,0,0,0)", borderwidth=0),
            margin        = dict(l=10, r=10, t=40, b=10),
            title_font_color = "#111827",
        )

        if PLOTLY and stats["por_dia"]:
            st.markdown("<div style='margin-top:1rem'></div>", unsafe_allow_html=True)
            _dd = pd.DataFrame(stats["por_dia"])
            _fig1 = px.bar(_dd, x="dia", y=["encontrados","total"], barmode="group",
                           labels={"dia":"Data","value":"Quantidade","variable":"Tipo"},
                           color_discrete_map={"encontrados":"#1D9E75","total":"rgba(55,138,221,0.5)"},
                           title="Buscas por dia — últimos 30 dias")
            _fig1.update_layout(**_layout_light)
            _fig1.update_traces(marker_line_width=0)
            st.plotly_chart(_fig1, use_container_width=True)

        if PLOTLY and stats["top_usuarios"]:
            _du = pd.DataFrame(stats["top_usuarios"])
            _fig2 = px.bar(_du, x="nome", y="execucoes",
                           labels={"nome":"Usuário","execucoes":"Execuções"},
                           color="execucoes",
                           color_continuous_scale=[[0,"rgba(13,110,86,0.3)"],[1,"#1D9E75"]],
                           title="Usuários mais ativos")
            _fig2.update_layout(**_layout_light, showlegend=False)
            _fig2.update_traces(marker_line_width=0)
            st.plotly_chart(_fig2, use_container_width=True)

        if not stats["por_dia"] and not stats["top_usuarios"]:
            st.markdown(
                """<div style="text-align:center;padding:3rem;color:#9ca3af;">
                    <div style="font-size:40px;margin-bottom:12px;">📊</div>
                    <div>Gráficos aparecerão após a primeira busca.</div>
                </div>""",
                unsafe_allow_html=True,
            )

    # ════════════════════════════════════════════════════════
    #  ABA 4 — CONFIGURAÇÕES
    # ════════════════════════════════════════════════════════
    with aba_config:
        secao_titulo("⚙️", "Configurações", "Conta, e-mail, cache e diagnóstico")

        _cc1, _cc2 = st.columns(2, gap="medium")

        with _cc1:
            st.markdown("#### 👤 Minha conta")
            with st.expander("Alterar senha"):
                _s1 = st.text_input("Senha atual",  type="password", key="s_atual")
                _s2 = st.text_input("Nova senha",   type="password", key="s_nova")
                _s3 = st.text_input("Confirmar",    type="password", key="s_conf")
                if st.button("💾 Salvar senha"):
                    from database import verificar_login as _vl
                    if not _vl(usuario.get("usuario", usuario.get("nome","")), _s1):
                        st.error("Senha atual incorreta.")
                    elif _s2 != _s3:   st.error("As senhas não coincidem.")
                    elif len(_s2) < 6: st.error("Mínimo 6 caracteres.")
                    else:
                        alterar_senha(usuario.get("id"), _s2)
                        st.success("✅ Senha alterada!")

            st.markdown("<div style='margin-top:1rem'></div>", unsafe_allow_html=True)
            st.markdown("#### 🗄️ Cache OCR")
            with st.expander("Gerenciar cache"):
                _cfs  = list(Path(__file__).parent.rglob(".ocr_*.json"))
                _csz  = sum(f.stat().st_size for f in _cfs if f.exists()) / 1024
                st.markdown(
                    f'<div style="display:flex;gap:24px;padding:8px 0;">'
                    f'<div><div style="font-size:22px;font-weight:700;color:#059669;">{len(_cfs)}</div>'
                    f'<div style="font-size:11px;color:#6b7280;">arquivos</div></div>'
                    f'<div><div style="font-size:22px;font-weight:700;color:#3b82f6;">{_csz:.0f} KB</div>'
                    f'<div style="font-size:11px;color:#6b7280;">tamanho</div></div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
                st.caption("Cache acelera re-execuções nos mesmos PDFs. Limpe se os PDFs foram modificados.")
                if st.button("🗑️ Limpar cache", key="btn_cache"):
                    _rm = sum(1 for f in _cfs if f.unlink(missing_ok=True) is None)
                    st.success(f"✅ {len(_cfs)} arquivo(s) removido(s).")
                    st.rerun()

        with _cc2:
            st.markdown("#### 📧 E-mail (SMTP)")
            with st.expander("Configurar servidor SMTP"):
                _smtp = cfg.get("smtp", {})
                _sh   = st.text_input("Servidor SMTP",  value=_smtp.get("host",""),       placeholder="smtp.gmail.com")
                _sp   = st.number_input("Porta",        value=int(_smtp.get("porta",587)), min_value=1, max_value=65535, step=1)
                _su   = st.text_input("Usuário/E-mail", value=_smtp.get("usuario",""))
                _ss   = st.text_input("Senha",          type="password", value=_smtp.get("senha",""))
                _sr   = st.text_input("Remetente",      value=_smtp.get("remetente",_smtp.get("usuario","")))
                st.caption("Gmail: smtp.gmail.com porta 587 + [Senha de App](https://myaccount.google.com/apppasswords)")
                _btn1, _btn2 = st.columns(2)
                with _btn1:
                    if st.button("💾 Salvar", use_container_width=True):
                        cfg["smtp"] = {"host":_sh,"porta":_sp,"usuario":_su,"senha":_ss,"remetente":_sr or _su}
                        salvar_config(cfg); st.success("✅ Salvo.")
                with _btn2:
                    if st.button("📤 Testar", use_container_width=True):
                        if not _sh or not _su or not _ss: st.error("Preencha os campos.")
                        else:
                            from mailer import enviar_email as _env
                            _ok, _err = _env(_su, "Teste — iFind", "<p>SMTP funcionando.</p>")
                            if _ok: st.success("✅ Enviado!")
                            else:   st.error(f"❌ {_err}")

        if usuario.get("admin"):
            st.markdown("#### 👥 Gerenciar usuários")
            _gu1, _gu2 = st.columns(2, gap="medium")
            with _gu1:
                with st.expander("Usuários cadastrados"):
                    _lu = listar_usuarios()
                    _dfu = pd.DataFrame(_lu)[["id","usuario","nome","email","admin","ativo","criado_em"]]
                    _dfu.columns = ["ID","Login","Nome","E-mail","Admin","Ativo","Criado em"]
                    st.dataframe(_dfu, use_container_width=True, hide_index=True)
            with _gu2:
                with st.expander("Criar novo usuário"):
                    _nu = st.text_input("Login",         key="nu")
                    _nn = st.text_input("Nome completo", key="nn")
                    _ne = st.text_input("E-mail",        key="ne")
                    _ns = st.text_input("Senha",         type="password", key="ns")
                    _na = st.checkbox("Administrador",   key="na")
                    if st.button("➕ Criar"):
                        if not _nu or not _nn or not _ns: st.error("Preencha login, nome e senha.")
                        elif len(_ns) < 6:                st.error("Senha: mínimo 6 caracteres.")
                        elif criar_usuario(_nu, _ns, _nn, _ne, int(_na)):
                            st.success(f"✅ Usuário '{_nu}' criado.")
                        else:
                            st.error(f"Login '{_nu}' já existe.")

        st.markdown("#### 🔬 Diagnóstico")
        with st.expander("Informações técnicas"):
            import platform as _pl
            _tp = shutil.which("tesseract") or "não encontrado no PATH"
            _pe = Path(__file__).parent / "tesseract_bin" / "tesseract.exe"
            if _pe.exists(): _tp += f" | portátil: {_pe}"
            st.code(
                f"Sistema     : {_pl.system()} {_pl.release()}\n"
                f"Python      : {_pl.python_version()}\n"
                f"Streamlit   : {st.__version__}\n"
                f"Tesseract   : {_tp}\n"
                f"Banco       : {(Path(__file__).parent / 'clinica.db').resolve()}\n"
                f"Config      : {CONFIG_PATH.resolve()}\n"
                f"Cache OCR   : {len(list(Path(__file__).parent.rglob('.ocr_*.json')))} arquivo(s)\n",
                language=None,
            )


_main()