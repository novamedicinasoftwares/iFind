"""
processor.py — Lógica de busca e extração de documentos PDF.

MELHORIAS desta versão:
  1. CAMINHOS INTELIGENTES:
     - Se o usuário já aponta para a pasta correta (ex: pasta do dia),
       o sistema detecta e usa ela diretamente sem tentar montar sub-caminhos.
     - Autodetecção de estrutura: drive/mês/dia, drive/dia, drive direto.
     - Sem duplicação de caminho no log.
     - Suporte a data somente com ano ou sem data (modo varredura).

  2. MODO VARREDURA TOTAL (novo):
     - Ativado quando a planilha tem só nome (sem data) ou o usuário marca
       "Buscar em todos os PDFs".
     - Varre TODOS os PDFs recursivamente a partir do drive_raiz.
     - Força DPI 100 automaticamente para agilizar.
     - Avisa o usuário que será lento antes de iniciar.

  3. DETECÇÃO DE NÍVEL DO CAMINHO:
     - Detecta automaticamente se o drive_raiz já é a pasta final,
       uma pasta de mês ou a raiz do drive.
     - Não monta sub-caminho se o destino já está resolvido.
"""

import fitz
import openpyxl
import unicodedata
import re
import io
import sys
import os
import threading
import hashlib
import json as _json
from pathlib import Path
from datetime import datetime, date
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    from rapidfuzz import fuzz as _fuzz
    _RAPIDFUZZ_OK = True
except ImportError:
    _fuzz = None
    _RAPIDFUZZ_OK = False

try:
    from PIL import Image as _PILImage
    import pytesseract as _pytesseract
    _OCR_OK = True
except ImportError:
    _PILImage = None
    _pytesseract = None
    _OCR_OK = False


# ---------------------------------------------------------------------------
# Cache OCR persistente
# ---------------------------------------------------------------------------

def _cache_pdf_path(caminho_pdf: Path) -> Path:
    return caminho_pdf.parent / f".ocr_{caminho_pdf.stem}.json"

def _pdf_hash(caminho_pdf: Path) -> str:
    try:
        with open(caminho_pdf, "rb") as f:
            return hashlib.md5(f.read(65536)).hexdigest()
    except Exception:
        return ""

def _cache_load(caminho_pdf: Path) -> dict:
    p = _cache_pdf_path(caminho_pdf)
    if not p.exists():
        return {}
    try:
        d = _json.loads(p.read_text(encoding="utf-8"))
        if d.get("hash") != _pdf_hash(caminho_pdf):
            p.unlink(missing_ok=True)
            return {}
        return d.get("paginas", {})
    except Exception:
        return {}

def _cache_save(caminho_pdf: Path, paginas: dict):
    try:
        _cache_pdf_path(caminho_pdf).write_text(
            _json.dumps({
                "hash": _pdf_hash(caminho_pdf),
                "gerado_em": datetime.now().isoformat(),
                "paginas": paginas,
            }, ensure_ascii=False),
            encoding="utf-8",
        )
    except Exception as e:
        print(f"[Cache] aviso ao salvar: {e}", file=sys.stderr)


# ---------------------------------------------------------------------------
# Utilitários de texto
# ---------------------------------------------------------------------------

def normalizar_texto(texto: str) -> str:
    if not texto:
        return ""
    texto = unicodedata.normalize("NFD", texto)
    texto = "".join(c for c in texto if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", texto.lower()).strip()


def nome_contem(texto_pagina: str, nome_buscado: str, threshold: float = 80.0) -> tuple[bool, float]:
    texto_norm = normalizar_texto(texto_pagina)
    nome_norm  = normalizar_texto(nome_buscado)

    palavras = nome_norm.split()
    if palavras and all(p in texto_norm for p in palavras):
        return True, 100.0

    if _RAPIDFUZZ_OK:
        palavras_texto = texto_norm.split()
        tam = len(palavras)
        melhor = 0.0
        for i in range(max(1, len(palavras_texto) - tam + 1)):
            janela = " ".join(palavras_texto[i: i + tam + 2])
            score  = _fuzz.token_set_ratio(nome_norm, janela)
            if score > melhor:
                melhor = score
        if melhor >= threshold:
            return True, float(melhor)

    return False, 0.0


# ---------------------------------------------------------------------------
# Seletor de pasta nativo
# ---------------------------------------------------------------------------

def abrir_seletor_pasta(titulo: str = "Selecionar pasta") -> str:
    import platform
    sistema = platform.system().lower()

    if sistema == "windows":
        try:
            import tkinter as tk
            from tkinter import filedialog
            root = tk.Tk()
            root.withdraw()
            root.wm_attributes("-topmost", True)
            pasta = filedialog.askdirectory(title=titulo)
            root.destroy()
            return pasta or ""
        except Exception as e:
            print(f"Erro ao abrir seletor Windows: {e}", file=sys.stderr)
            return ""
    elif sistema == "darwin":
        try:
            import subprocess
            script = f'choose folder with prompt "{titulo}"'
            r = subprocess.run(["osascript", "-e", script], capture_output=True, text=True)
            if r.returncode == 0:
                caminho = r.stdout.strip().replace(":", "/")
                return caminho if caminho.startswith("/") else "/" + caminho
        except Exception:
            return ""
    else:
        for cmd in [["zenity", "--file-selection", "--directory", f"--title={titulo}"],
                    ["yad", "--file", "--directory", f"--title={titulo}"]]:
            try:
                import subprocess
                import shutil
                if not shutil.which(cmd[0]):
                    continue
                r = subprocess.run(cmd, capture_output=True, text=True)
                if r.returncode == 0:
                    return r.stdout.strip()
            except Exception:
                continue
    return ""


# ---------------------------------------------------------------------------
# Leitura inteligente da planilha Excel
# ---------------------------------------------------------------------------

_PALAVRAS_NOME  = {"nome", "paciente", "patient", "name", "beneficiario", "segurado", "cliente", "pessoa", "titular"}
_PALAVRAS_DATA  = {"data", "date", "dt", "dia", "procedimento", "proc", "atendimento", "consulta", "exame", "realizacao", "competencia", "ano"}
_PALAVRAS_EMAIL = {"email", "e-mail", "correio", "mail"}
_PALAVRAS_LIXO  = {"total", "subtotal", "soma", "media", "count", "grand total", "totais", "#", "n/a", "n.a."}

def _celula_para_str(valor) -> str:
    if valor is None:
        return ""
    if isinstance(valor, (datetime, date)):
        return str(valor)
    return str(valor).strip()

def _parece_data(valor) -> bool:
    if isinstance(valor, (datetime, date)):
        return True
    s = str(valor).strip()
    padroes = [
        r"^\d{1,2}/\d{1,2}(/\d{2,4})?$",
        r"^\d{4}-\d{2}-\d{2}",
        r"^\d{1,2}-\d{1,2}-\d{2,4}$",
        r"^\d{4}$",   # só ano
    ]
    return any(re.match(p, s) for p in padroes)

def _parece_email(valor) -> bool:
    s = str(valor).strip() if valor else ""
    return bool(re.match(r"[^@]+@[^@]+\.[^@]+", s))

def _parece_nome(valor) -> bool:
    s = str(valor).strip() if valor else ""
    palavras = [p for p in s.split() if re.search(r"[a-zA-ZÀ-ú]", p)]
    return len(palavras) >= 2 and not _parece_data(valor) and not _parece_email(valor)

def _linha_e_lixo(linha: tuple) -> bool:
    valores = [_celula_para_str(v).lower() for v in linha if v is not None]
    if not valores:
        return True
    for v in valores:
        if any(lixo in v for lixo in _PALAVRAS_LIXO):
            return True
    return False

def _detectar_coluna(cabecalho: list[str], palavras_chave: set, fallback: int) -> int:
    for i, h in enumerate(cabecalho):
        h_norm = normalizar_texto(h)
        if any(p in h_norm for p in palavras_chave):
            return i
    return fallback

def _score_celula_cabecalho(celula_str: str) -> int:
    s = celula_str.lower().strip()
    if not s:
        return 0
    score = 0
    if s in _PALAVRAS_NOME:  score += 4
    if s in _PALAVRAS_DATA:  score += 4
    if s in _PALAVRAS_EMAIL: score += 2
    return score

def _encontrar_linha_cabecalho(ws) -> tuple[int, bool]:
    melhor_linha, melhor_score = 1, 0
    for num_linha in range(1, min(11, ws.max_row + 1)):
        score = sum(_score_celula_cabecalho(_celula_para_str(c.value)) for c in ws[num_linha])
        if score > melhor_score:
            melhor_score, melhor_linha = score, num_linha
    return melhor_linha, (melhor_score >= 4)

def _detectar_colunas_por_conteudo(ws, primeira_linha: int) -> tuple[int, int, int]:
    num_amostras = min(primeira_linha + 15, ws.max_row + 1) - primeira_linha
    if num_amostras <= 0:
        return 0, 1, -1

    score_nome  = {}
    score_data  = {}
    score_email = {}

    for num_linha in range(primeira_linha, primeira_linha + num_amostras):
        if num_linha > ws.max_row:
            break
        for col_idx, cell in enumerate(ws[num_linha]):
            v = cell.value
            if v is None:
                continue
            if _parece_nome(v):
                score_nome[col_idx]  = score_nome.get(col_idx, 0) + 1
            if _parece_data(v):
                score_data[col_idx]  = score_data.get(col_idx, 0) + 1
            if _parece_email(v):
                score_email[col_idx] = score_email.get(col_idx, 0) + 1

    col_nome  = max(score_nome,  key=score_nome.get)  if score_nome  else 0
    col_data  = max(score_data,  key=score_data.get)  if score_data  else 1
    col_email = max(score_email, key=score_email.get) if score_email else -1

    if col_nome == col_data:
        candidatos_data = sorted(score_data, key=score_data.get, reverse=True)
        col_data = next((c for c in candidatos_data if c != col_nome), col_nome + 1)

    return col_nome, col_data, col_email

def _ler_aba(ws) -> list[dict]:
    if ws.max_row < 1:
        return []
    linha_cab, tem_cabecalho = _encontrar_linha_cabecalho(ws)
    if tem_cabecalho:
        cabecalho = [_celula_para_str(c.value) for c in ws[linha_cab]]
        col_nome  = _detectar_coluna(cabecalho, _PALAVRAS_NOME,  0)
        col_data  = _detectar_coluna(cabecalho, _PALAVRAS_DATA,  1)
        col_email = _detectar_coluna(cabecalho, _PALAVRAS_EMAIL, -1)
        primeira_linha_dados = linha_cab + 1
        nomes_cabecalho = {normalizar_texto(h) for h in cabecalho if h}
    else:
        col_nome, col_data, col_email = _detectar_colunas_por_conteudo(ws, 1)
        primeira_linha_dados = 1
        nomes_cabecalho = set()

    registros = []
    for num_linha in range(primeira_linha_dados, ws.max_row + 1):
        linha = tuple(c.value for c in ws[num_linha])
        if all(v is None for v in linha) or _linha_e_lixo(linha):
            continue
        nome = _celula_para_str(linha[col_nome])
        if not nome or normalizar_texto(nome) in nomes_cabecalho:
            continue
        data  = linha[col_data]  if col_data  < len(linha) else None
        email = _celula_para_str(linha[col_email]) if 0 <= col_email < len(linha) else ""
        registros.append({"nome": nome, "data": data, "email": email})
    return registros

def ler_planilha(caminho_excel: str) -> list[dict]:
    wb = openpyxl.load_workbook(caminho_excel, data_only=True)
    melhor_registros = []
    for ws in [wb.active] + [wb[n] for n in wb.sheetnames if wb[n] != wb.active]:
        try:
            regs = _ler_aba(ws)
            if len(regs) > len(melhor_registros):
                melhor_registros = regs
        except Exception:
            continue
    if not melhor_registros:
        raise ValueError("Planilha vazia ou ilegível.")
    return melhor_registros


def validar_planilha(registros: list[dict]) -> list[dict]:
    """Detecta duplicatas, datas ausentes e outros avisos."""
    avisos = []
    vistos: dict[tuple, int] = {}
    sem_data = 0
    for i, reg in enumerate(registros, 1):
        nome = reg.get("nome", "").strip()
        data = str(reg.get("data", "")).strip()
        if len(nome.split()) < 2:
            avisos.append({"tipo": "nome_curto",
                           "msg": f"Nome muito curto na linha {i}: '{nome}'", "linha": i})
        chave = (normalizar_texto(nome), data)
        if chave in vistos:
            avisos.append({"tipo": "duplicata",
                           "msg": f"Duplicata: '{nome}' em {data} (linhas {vistos[chave]} e {i})",
                           "linha": i})
        else:
            vistos[chave] = i
        if not data or data in ("None", ""):
            sem_data += 1
    if sem_data > 0:
        avisos.append({"tipo": "sem_data",
                       "msg": f"{sem_data} registro(s) sem data — será usado modo varredura total.",
                       "linha": 0})
    return avisos


# ---------------------------------------------------------------------------
# Navegação e Caminhos — VERSÃO INTELIGENTE
# ---------------------------------------------------------------------------

def extrair_componentes_data(data) -> dict:
    """
    Extrai dia, mês e ano de qualquer formato de data.
    Retorna dict com chaves: dia, mes, ano, tem_dia, tem_mes, tem_ano.
    """
    resultado = {"dia": None, "mes": None, "ano": None,
                 "tem_dia": False, "tem_mes": False, "tem_ano": False}

    if data is None or str(data).strip() in ("", "None"):
        return resultado

    if isinstance(data, (datetime, date)):
        resultado.update({
            "dia": data.strftime("%d"), "mes": data.strftime("%m"),
            "ano": str(data.year),
            "tem_dia": True, "tem_mes": True, "tem_ano": True,
        })
        return resultado

    texto = str(data).strip()

    # Só ano: "2025" ou "2024"
    if re.match(r"^\d{4}$", texto):
        resultado.update({"ano": texto, "tem_ano": True})
        return resultado

    # DD/MM/YYYY ou DD/MM/YY ou DD/MM
    m = re.match(r"^(\d{1,2})/(\d{1,2})(?:/(\d{2,4}))?$", texto)
    if m:
        resultado.update({
            "dia": m.group(1).zfill(2), "mes": m.group(2).zfill(2),
            "tem_dia": True, "tem_mes": True,
        })
        if m.group(3):
            ano = m.group(3)
            resultado["ano"] = "20" + ano if len(ano) == 2 else ano
            resultado["tem_ano"] = True
        return resultado

    # YYYY-MM-DD
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", texto)
    if m:
        resultado.update({
            "dia": m.group(3), "mes": m.group(2), "ano": m.group(1),
            "tem_dia": True, "tem_mes": True, "tem_ano": True,
        })
        return resultado

    # MM/YYYY ou MM-YYYY (sem dia)
    m = re.match(r"^(\d{1,2})[/-](\d{4})$", texto)
    if m:
        resultado.update({
            "mes": m.group(1).zfill(2), "ano": m.group(2),
            "tem_mes": True, "tem_ano": True,
        })
        return resultado

    # Fallback: só ano se vier algo tipo "2025-..."
    m = re.match(r"^(\d{4})", texto)
    if m:
        resultado.update({"ano": m.group(1), "tem_ano": True})
    return resultado


# Mapa de nomes padrão de mês
MESES = {
    "01": "1.JANEIRO",  "02": "2.FEVEREIRO", "03": "3.MARCO",
    "04": "4.ABRIL",    "05": "5.MAIO",       "06": "6.JUNHO",
    "07": "7.JULHO",    "08": "8.AGOSTO",     "09": "9.SETEMBRO",
    "10": "10.OUTUBRO", "11": "11.NOVEMBRO",  "12": "12.DEZEMBRO",
}

_NOMES_MESES_PT = {
    "janeiro": "01", "jan": "01",
    "fevereiro": "02", "fev": "02",
    "marco": "03", "março": "03", "mar": "03",
    "abril": "04", "abr": "04",
    "maio": "05", "mai": "05",
    "junho": "06", "jun": "06",
    "julho": "07", "jul": "07",
    "agosto": "08", "ago": "08",
    "setembro": "09", "set": "09",
    "outubro": "10", "out": "10",
    "novembro": "11", "nov": "11",
    "dezembro": "12", "dez": "12",
}

_CACHE_MESES: dict[str, dict[str, str]] = {}


def autodetectar_meses(drive_raiz: str) -> dict[str, str]:
    """Detecta nomes reais das pastas de mês no drive. Cache por sessão."""
    if drive_raiz in _CACHE_MESES:
        return _CACHE_MESES[drive_raiz]

    raiz = Path(drive_raiz)
    if not raiz.exists():
        return {}

    mapa: dict[str, str] = {}
    for pasta in raiz.iterdir():
        if not pasta.is_dir():
            continue
        nome_lower = normalizar_texto(pasta.name)
        for palavra, num in _NOMES_MESES_PT.items():
            if palavra in nome_lower:
                existente = mapa.get(num)
                if existente is None:
                    mapa[num] = pasta.name
                else:
                    try:
                        if len(list((raiz / pasta.name).iterdir())) > len(list((raiz / existente).iterdir())):
                            mapa[num] = pasta.name
                    except Exception:
                        pass
                break

    _CACHE_MESES[drive_raiz] = mapa
    return mapa


def _detectar_nivel_pasta(pasta: Path) -> str:
    """
    Detecta o nível da pasta informada pelo usuário:
      'dia'  — já contém PDFs diretamente (pasta final)
      'mes'  — contém subpastas de dia
      'raiz' — contém subpastas de mês
      'vazio'— não existe ou vazia
    """
    if not pasta.exists():
        return "vazio"
    subpastas = [p for p in pasta.iterdir() if p.is_dir()]
    pdfs_diretos = list(pasta.glob("*.pdf"))

    if pdfs_diretos:
        return "dia"

    # Verifica se subpastas parecem dias (ex: "15.01", "02.03")
    padroes_dia = [re.match(r"^\d{1,2}[.\-]\d{2}$", p.name) for p in subpastas]
    if any(padroes_dia):
        return "mes"

    # Verifica se subpastas parecem meses
    for p in subpastas:
        nome_lower = normalizar_texto(p.name)
        if any(palavra in nome_lower for palavra in _NOMES_MESES_PT):
            return "raiz"

    # Se tem subpastas mas não identificou — trata como raiz
    if subpastas:
        return "raiz"

    return "vazio"


def resolver_pasta_pdfs(drive_raiz: str, data) -> tuple[Path, str]:
    """
    Resolve o caminho final onde estão os PDFs, dado o drive_raiz e a data.

    Lógica inteligente em cascata:
      1. Se drive_raiz já contém PDFs → usa direto (usuário apontou para pasta final)
      2. Se drive_raiz é pasta de mês → monta pasta/DD.MM
      3. Se drive_raiz é raiz do drive → monta pasta/MÊS/DD.MM (comportamento original)
      4. Se data é só ano → retorna pasta do ano se encontrar, ou drive_raiz
      5. Se sem data → retorna drive_raiz (será usado varredura)

    Retorna (Path resolvido, descrição do que foi feito para o log).
    """
    raiz = Path(drive_raiz)
    nivel = _detectar_nivel_pasta(raiz)
    comp  = extrair_componentes_data(data)

    # ── Caso 1: pasta já tem PDFs — usa direto ──────────────────
    if nivel == "dia":
        return raiz, f"pasta direta (PDFs encontrados em {raiz.name})"

    # ── Sem data ou só sem componentes úteis — varredura ────────
    if not comp["tem_dia"] and not comp["tem_mes"] and not comp["tem_ano"]:
        return raiz, "sem data — varredura"

    # ── Caso: só ano informado ───────────────────────────────────
    if comp["tem_ano"] and not comp["tem_mes"]:
        # Procura pasta com o ano no nome dentro do drive_raiz
        ano = comp["ano"]
        for p in raiz.iterdir() if raiz.exists() else []:
            if p.is_dir() and ano in p.name:
                return p, f"pasta do ano {ano} detectada: {p.name}"
        return raiz, f"ano {ano} — varredura na raiz"

    # ── Caso: tem mês mas não tem dia ────────────────────────────
    if comp["tem_mes"] and not comp["tem_dia"]:
        mes_num = comp["mes"]
        if nivel == "raiz":
            mapa = autodetectar_meses(drive_raiz)
            nome_mes = mapa.get(mes_num, MESES.get(mes_num, mes_num))
            pasta_mes = raiz / nome_mes
            return pasta_mes, f"pasta do mês: {nome_mes}"
        # Se nível for mes, já está na pasta certa
        return raiz, f"pasta do mês (já apontada)"

    # ── Caso principal: tem dia e mês ────────────────────────────
    dia    = comp["dia"]
    mes    = comp["mes"]
    pasta_dia_nome = f"{dia}.{mes}"

    if nivel == "dia":
        # Já está na pasta certa — não adiciona sub-caminho
        return raiz, f"pasta direta usada: {raiz.name}"

    if nivel == "mes":
        # drive_raiz é pasta de mês — só adiciona o dia
        pasta_final = raiz / pasta_dia_nome
        return pasta_final, f"pasta do mês/{pasta_dia_nome}"

    if nivel == "raiz":
        # drive_raiz é raiz — monta mês + dia
        mapa = autodetectar_meses(drive_raiz)
        nome_mes = mapa.get(mes, MESES.get(mes, mes))
        pasta_final = raiz / nome_mes / pasta_dia_nome
        return pasta_final, f"{nome_mes}/{pasta_dia_nome}"

    # fallback
    return raiz / pasta_dia_nome, pasta_dia_nome


# Mantém compatibilidade com o app.py que chama montar_caminho_pasta
def montar_caminho_pasta(drive_raiz: str, data) -> Path:
    pasta, _ = resolver_pasta_pdfs(drive_raiz, data)
    return pasta


# ---------------------------------------------------------------------------
# Detecção de modo de busca por registro
# ---------------------------------------------------------------------------

def classificar_modo_registro(data) -> str:
    """
    Classifica o modo de busca para um registro:
      'dia'      — tem dia+mês (pode ir direto para a pasta)
      'mes'      — tem só mês (busca na pasta do mês inteiro)
      'ano'      — tem só ano (busca na pasta do ano)
      'varredura'— sem data (busca em tudo)
    """
    comp = extrair_componentes_data(data)
    if comp["tem_dia"] and comp["tem_mes"]:
        return "dia"
    if comp["tem_mes"]:
        return "mes"
    if comp["tem_ano"]:
        return "ano"
    return "varredura"


def coletar_pdfs_recursivo(pasta_raiz: Path, limite: int = 5000) -> list[Path]:
    """
    Coleta todos os PDFs recursivamente a partir de pasta_raiz.
    Limite de segurança para não travar em drives enormes.
    """
    pdfs = []
    try:
        for pdf in sorted(pasta_raiz.rglob("*.pdf")):
            pdfs.append(pdf)
            if len(pdfs) >= limite:
                break
    except Exception:
        pass
    return pdfs


# ---------------------------------------------------------------------------
# Configuração do Tesseract
# ---------------------------------------------------------------------------

def _configurar_tesseract() -> tuple[str, str]:
    import pytesseract

    tessdata_dir = ""
    lang = "por"

    cfg_path = Path(__file__).parent / "config_tesseract.py"
    if cfg_path.exists():
        try:
            exec(cfg_path.read_text(encoding="utf-8"), {"pytesseract": pytesseract})
        except Exception as e:
            print(f"[Tesseract] Aviso ao ler config_tesseract.py: {e}", file=sys.stderr)

    cmd_path = pytesseract.pytesseract.tesseract_cmd
    if cmd_path and os.path.exists(str(cmd_path)):
        pasta_bin = Path(cmd_path).parent
        for candidato in [pasta_bin / "tessdata",
                          pasta_bin.parent / "tessdata",
                          pasta_bin / "share" / "tessdata"]:
            if candidato.exists():
                tessdata_dir = str(candidato.resolve())
                break

    if not tessdata_dir:
        tessdata_dir = os.environ.get("TESSDATA_PREFIX", "")

    if tessdata_dir:
        os.environ["TESSDATA_PREFIX"] = tessdata_dir

    try:
        langs_disp = pytesseract.get_languages(config="")
        if "por" not in langs_disp:
            lang = "eng"
    except Exception:
        try:
            from PIL import Image as _I
            pytesseract.image_to_string(_I.new("RGB", (10, 10), "white"), lang="por")
        except Exception:
            lang = "eng"

    if lang == "eng":
        print("[Tesseract] Usando lang=eng (por.traineddata não disponível).", file=sys.stderr)

    return tessdata_dir, lang


_TESS_CONFIG_CACHE: dict = {}

def _get_tess_config() -> tuple[str, str]:
    if not _TESS_CONFIG_CACHE:
        tessdata_dir, lang = _configurar_tesseract()
        _TESS_CONFIG_CACHE["tessdata_dir"] = tessdata_dir
        _TESS_CONFIG_CACHE["lang"] = lang
    return _TESS_CONFIG_CACHE["tessdata_dir"], _TESS_CONFIG_CACHE["lang"]


# ---------------------------------------------------------------------------
# Modos de extração
# ---------------------------------------------------------------------------

MODO_AUTO   = "auto"
MODO_NATIVO = "nativo"
MODO_OCR    = "ocr"


# ---------------------------------------------------------------------------
# Filtro ASO
# ---------------------------------------------------------------------------

_ASO_MARCADORES_INCLUSAO = {
    "aso", "atestado de saude ocupacional", "admissional", "demissional",
    "retorno ao trabalho", "mudanca de riscos ocupacionais", "periodico",
    "nr-7", "pcmso", "exame clinico", "riscos", "inapto", "apto",
    "medico responsavel",
}

_ASO_MARCADORES_EXCLUSAO = {
    "prontuario de pericia medica", "prontuario", "pericia medica",
    "cid:", "dias de afastamento", "data do atestado",
    "afastamento total", "afastamento parcial", "receita medica",
    "prescricao medica", "atestado medico", "laudo medico",
    "relatorio medico",
}

_ASO_SCORE_MINIMO = 2

def _pagina_e_aso(texto: str) -> bool:
    texto_norm = normalizar_texto(texto)
    for excluir in _ASO_MARCADORES_EXCLUSAO:
        if excluir in texto_norm:
            return False
    score = sum(1 for inc in _ASO_MARCADORES_INCLUSAO if inc in texto_norm)
    return score >= _ASO_SCORE_MINIMO


# ---------------------------------------------------------------------------
# Extração de texto por página
# ---------------------------------------------------------------------------

def extrair_texto_pagina(pagina: fitz.Page, forcar_ocr: bool = False, dpi: int = 150) -> str:
    if not forcar_ocr:
        texto = pagina.get_text()
        if len(texto.strip()) >= 50:
            return texto
    else:
        texto = ""

    if not _OCR_OK:
        return texto

    try:
        _, lang = _get_tess_config()
        pix = pagina.get_pixmap(dpi=dpi)
        img = _PILImage.open(io.BytesIO(pix.tobytes("png")))
        texto_ocr = _pytesseract.image_to_string(img, lang=lang)
        return texto + "\n" + texto_ocr
    except Exception as e:
        print(f"[OCR] Erro na página: {e}", file=sys.stderr)
        return texto


def sanitizar_nome_arquivo(nome: str) -> str:
    sem_acento = unicodedata.normalize("NFD", nome)
    sem_acento = "".join(c for c in sem_acento if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", "_", re.sub(r"[^\w\s-]", "", sem_acento).strip())


def extrair_e_salvar_paginas(
    caminho_pdf: Path,
    paginas_scores: list[tuple[int, float]],
    pasta_destino: Path,
    nome_paciente: str,
) -> Path:
    pasta_destino.mkdir(parents=True, exist_ok=True)
    nome_base = sanitizar_nome_arquivo(nome_paciente)
    saida = pasta_destino / f"{nome_base}.pdf"
    contador = 2
    while saida.exists():
        saida = pasta_destino / f"{nome_base}_{contador}.pdf"
        contador += 1
    doc_orig = fitz.open(str(caminho_pdf))
    doc_novo = fitz.open()
    for idx, _ in sorted(paginas_scores):
        doc_novo.insert_pdf(doc_orig, from_page=idx, to_page=idx)
    doc_novo.save(str(saida))
    doc_novo.close()
    doc_orig.close()
    return saida


# ---------------------------------------------------------------------------
# Worker de PDF
# ---------------------------------------------------------------------------

def _pdf_e_digital(doc: fitz.Document, amostras: int = 5) -> bool:
    total = len(doc)
    if total == 0:
        return False
    indices = [int(i * total / min(amostras, total)) for i in range(min(amostras, total))]
    com_texto = sum(1 for i in indices if len(doc[i].get_text().strip()) >= 50)
    return com_texto >= max(1, len(indices) // 2)


def _worker_pdf(
    j: int,
    pdf: Path,
    nomes: list[str],
    threshold: float,
    modo: str,
    dpi: int,
    n_pdfs: int,
    filtrar_aso: bool = False,
) -> tuple[int, Path, dict[str, tuple[list, float, int]], list[str]]:
    logs: list[str] = []
    cache = _cache_load(pdf)
    dirty = False
    textos: dict[int, str] = {}
    textos_descartados: dict[int, str] = {}

    try:
        doc = fitz.open(str(pdf))
    except Exception as e:
        logs.append(f"ERRO ao abrir {pdf.name}: {e}")
        return j, pdf, {}, logs

    total_pags = len(doc)
    modo_efetivo = modo
    if modo == MODO_AUTO:
        eh_digital = _pdf_e_digital(doc)
        modo_efetivo = MODO_NATIVO if eh_digital else MODO_OCR
        logs.append(
            f"PDF {j+1}/{n_pdfs}  tipo: {'digital' if eh_digital else 'escaneado'}"
            f" → modo {modo_efetivo}  {pdf.name}"
        )

    nomes_pendentes = set(nomes)

    for num, pagina in enumerate(doc):
        if not nomes_pendentes:
            logs.append(
                f"PDF {j+1}/{n_pdfs}  early-exit pág {num+1} "
                f"(todos encontrados)  {pdf.name}"
            )
            break

        chave = f"{num}_{dpi}_{modo_efetivo}"

        if chave in cache:
            texto = cache[chave]
            logs.append(f"PDF {j+1}/{n_pdfs}  pág {num+1}/{total_pags} (cache)  {pdf.name}")
        else:
            if modo_efetivo == MODO_NATIVO:
                logs.append(f"PDF {j+1}/{n_pdfs}  pág {num+1}/{total_pags} (nativo)  {pdf.name}")
                texto = pagina.get_text()
            else:
                logs.append(f"PDF {j+1}/{n_pdfs}  pág {num+1}/{total_pags} (OCR)  {pdf.name}")
                texto = extrair_texto_pagina(pagina, forcar_ocr=True, dpi=dpi)

            if texto.strip():
                cache[chave] = texto
                dirty = True

        if filtrar_aso and not _pagina_e_aso(texto):
            textos_descartados[num] = texto
        else:
            textos[num] = texto
            for nome in list(nomes_pendentes):
                achou, score = nome_contem(texto, nome, threshold)
                if achou and score >= threshold:
                    nomes_pendentes.discard(nome)

    doc.close()

    if dirty:
        _cache_save(pdf, cache)

    resultados: dict[str, tuple[list, float, int]] = {}
    for nome in nomes:
        encontradas: list[tuple[int, float]] = []
        melhor_score = 0.0
        achou_em_descartada = 0

        for num, texto in textos.items():
            achou, score = nome_contem(texto, nome, threshold)
            if score > melhor_score:
                melhor_score = score
            if achou:
                encontradas.append((num, score))
                if score >= 100.0:
                    break

        if filtrar_aso:
            for num, texto in textos_descartados.items():
                achou, _ = nome_contem(texto, nome, threshold)
                if achou:
                    achou_em_descartada += 1

        if encontradas:
            encontradas.sort(key=lambda x: x[1], reverse=True)
        resultados[nome] = (encontradas, melhor_score, achou_em_descartada)

    logs.append(f"PDF {j+1}/{n_pdfs} concluído  {pdf.name}")
    return j, pdf, resultados, logs


# ---------------------------------------------------------------------------
# Função Principal
# ---------------------------------------------------------------------------

def processar_lista(
    caminho_excel: str,
    drive_raiz: str,
    pasta_destino: str,
    threshold_fuzzy: float = 80.0,
    callback=None,
    modo_extracao: str = MODO_AUTO,
    dpi_ocr: int = 150,
    max_workers: int = 4,
    filtrar_aso: bool = False,
    varredura_total: bool = False,
) -> list[dict]:
    """
    Processa todos os registros da planilha.

    Parâmetro varredura_total:
        Se True, ignora as datas e varre TODOS os PDFs recursivamente
        a partir do drive_raiz. Útil quando só se tem o nome.
        DPI é forçado para 100 automaticamente para agilizar.
    """
    def _cb(prog: float, etapa: str, detalhe: str = "", status: str = "info"):
        if callback:
            callback(prog, etapa, detalhe, status)

    _CACHE_MESES.clear()

    _cb(0.0, "Lendo planilha", "Aguarde...")
    registros = ler_planilha(caminho_excel)
    total = len(registros)
    _cb(0.0, "Planilha lida", f"{total} registros encontrados", "ok")

    # ── Detecta se precisa de varredura total ────────────────────
    sem_data = sum(1 for r in registros
                   if not r.get("data") or str(r["data"]).strip() in ("", "None"))
    modo_varredura = varredura_total or (sem_data == total)

    if modo_varredura:
        dpi_ocr = 100  # força DPI baixo para agilizar
        _cb(0.0, "Modo varredura total ativado",
            f"DPI forçado para 100 — buscando em todos os PDFs de {drive_raiz}", "aviso")

    # ── Agrupa registros por pasta resolvida ─────────────────────
    grupos: dict[str, list[dict]] = {}

    for reg in registros:
        if modo_varredura:
            chave = drive_raiz
        else:
            modo_reg = classificar_modo_registro(reg.get("data"))
            if modo_reg == "varredura":
                chave = drive_raiz
            else:
                try:
                    pasta_resolvida, descricao = resolver_pasta_pdfs(drive_raiz, reg["data"])
                    chave = str(pasta_resolvida)
                except Exception:
                    chave = "__erro__"
        grupos.setdefault(chave, []).append(reg)

    n_grupos = len(grupos)
    _cb(0.0, "Agrupamento concluído",
        f"{n_grupos} pasta(s) distintas para {total} registro(s)", "ok")

    # Mapa de resultados
    resultados_map: dict[str, dict] = {}
    for reg in registros:
        resultados_map[reg["nome"]] = {
            "nome": reg["nome"], "data": str(reg.get("data", "")),
            "email": reg.get("email", ""),
            "encontrado": False, "arquivo": "", "erro": "", "score_fuzzy": 0.0,
        }

    # Retomada automática
    destino_path = Path(pasta_destino)
    pulados = 0
    for reg in registros:
        nome = reg["nome"]
        arq = destino_path / (sanitizar_nome_arquivo(nome) + ".pdf")
        if arq.exists():
            resultados_map[nome].update({
                "encontrado": True, "arquivo": str(arq), "score_fuzzy": 100.0,
            })
            pulados += 1

    if pulados > 0:
        _cb(0.0, "Retomando execução",
            f"{pulados} já encontrado(s) na pasta de destino — pulando", "ok")

    grupos_feitos = 0

    for chave, regs_grupo_orig in grupos.items():
        grupos_feitos += 1
        prog_base = (grupos_feitos - 1) / n_grupos
        prog_fim  = grupos_feitos       / n_grupos

        regs_grupo  = [r for r in regs_grupo_orig
                       if not resultados_map[r["nome"]]["encontrado"]]
        nomes_grupo = [r["nome"] for r in regs_grupo]
        n_nomes     = len(nomes_grupo)

        if n_nomes == 0:
            _cb(prog_fim, "Grupo já concluído", Path(chave).name, "ok")
            continue

        if chave == "__erro__":
            for r in regs_grupo:
                resultados_map[r["nome"]]["erro"] = "Erro ao montar caminho da pasta."
            _cb(prog_fim, "Erro de caminho", "", "erro")
            continue

        pasta = Path(chave)

        # ── Descrição limpa para o log (sem duplicar o drive_raiz) ──
        try:
            descricao_pasta = pasta.relative_to(drive_raiz).as_posix()
        except ValueError:
            descricao_pasta = pasta.name

        _cb(prog_base, "Verificando pasta",
            f"{descricao_pasta}  —  {n_nomes} paciente(s)")

        if not pasta.exists():
            for r in regs_grupo:
                resultados_map[r["nome"]]["erro"] = (
                    f"Pasta não encontrada: {descricao_pasta}"
                )
            _cb(prog_fim, "Pasta não encontrada", descricao_pasta, "erro")
            continue

        # ── Coleta PDFs ──────────────────────────────────────────
        if modo_varredura or chave == drive_raiz:
            pdfs = coletar_pdfs_recursivo(pasta)
            _cb(prog_base, "Varredura recursiva",
                f"{len(pdfs)} PDF(s) encontrados em {descricao_pasta}", "info")
        else:
            pdfs = sorted(pasta.glob("*.pdf"))

        n_pdfs = len(pdfs)

        if n_pdfs == 0:
            for r in regs_grupo:
                resultados_map[r["nome"]]["erro"] = f"Nenhum PDF em: {descricao_pasta}"
            _cb(prog_fim, "Sem PDFs", descricao_pasta, "aviso")
            continue

        _cb(prog_base, "Pasta localizada",
            f"{n_pdfs} PDF(s)  —  buscando {n_nomes} nome(s)", "ok")

        melhor: dict[str, tuple] = {}
        lock   = threading.Lock()
        parar  = threading.Event()

        workers = min(max_workers, n_pdfs)
        with ThreadPoolExecutor(max_workers=workers) as ex:
            futuros = {
                ex.submit(
                    _worker_pdf,
                    j, pdf,
                    list(nomes_grupo),
                    threshold_fuzzy,
                    modo_extracao,
                    dpi_ocr,
                    n_pdfs,
                    filtrar_aso,
                ): (j, pdf)
                for j, pdf in enumerate(pdfs)
            }

            for fut in as_completed(futuros):
                if parar.is_set():
                    break

                try:
                    j, pdf, res, logs = fut.result()
                except Exception as e:
                    import traceback as _tb
                    _cb(prog_base, "Erro no worker",
                        _tb.format_exc().strip().splitlines()[-1], "erro")
                    continue

                prog_pdf = prog_base + (prog_fim - prog_base) * ((j + 1) / n_pdfs)
                for linha in logs:
                    _cb(prog_pdf, linha, "", "info")

                with lock:
                    for nome, (pags, score, em_descartada) in res.items():
                        entrada_atual = melhor.get(nome, (None, [], 0.0, 0))
                        sc_atual = entrada_atual[2]
                        desc_atual = entrada_atual[3]
                        if score > sc_atual:
                            melhor[nome] = (pdf, pags, score, em_descartada)
                        elif em_descartada > desc_atual:
                            melhor[nome] = (entrada_atual[0], entrada_atual[1],
                                            entrada_atual[2], entrada_atual[3] + em_descartada)

                    todos_ok = all(
                        melhor.get(n, (None, [], 0.0, 0))[2] >= threshold_fuzzy
                        for n in nomes_grupo
                    )

                if todos_ok:
                    parar.set()
                    for f in futuros:
                        f.cancel()
                    break

        # Salva resultados do grupo
        enc_grupo = 0
        for reg in regs_grupo:
            nome = reg["nome"]
            entrada = melhor.get(nome, (None, [], 0.0, 0))
            pdf_src, pags, score, em_descartada = entrada

            if pags:
                try:
                    saida = extrair_e_salvar_paginas(
                        pdf_src, pags, Path(pasta_destino), nome
                    )
                    resultados_map[nome].update({
                        "encontrado": True,
                        "arquivo": str(saida),
                        "score_fuzzy": score,
                    })
                    _cb(prog_fim, "PDF salvo",
                        f"{nome}  score {score:.0f}%", "pdf_salvo")
                    enc_grupo += 1
                except Exception as e:
                    resultados_map[nome]["erro"] = f"Erro ao salvar: {e}"
                    _cb(prog_fim, "Erro ao salvar", str(e), "erro")
            else:
                if filtrar_aso and em_descartada > 0:
                    msg_erro = (
                        f"Nome encontrado em {em_descartada} página(s), "
                        f"mas nenhuma é ASO. Desative o filtro ASO para extrair."
                    )
                    resultados_map[nome]["erro"] = msg_erro
                    _cb(prog_fim, "Filtrado — não é ASO",
                        f"{nome}  —  {em_descartada} pág(s) descartada(s)", "aviso")
                else:
                    detalhe = f"Buscado em {n_pdfs} PDF(s)"
                    if score > 0:
                        detalhe += f" — melhor score: {score:.0f}% (threshold: {threshold_fuzzy:.0f}%)"
                    resultados_map[nome]["erro"] = "Nome não localizado nos PDFs."
                    _cb(prog_fim, "Não encontrado", detalhe, "aviso")

        _cb(prog_fim, "Pasta concluída",
            f"{descricao_pasta}  —  {enc_grupo}/{n_nomes} encontrados",
            "ok" if enc_grupo == n_nomes else "aviso")

    resultados = [resultados_map[r["nome"]] for r in registros]
    enc_total  = sum(r["encontrado"] for r in resultados)
    _cb(1.0, "Concluído", f"{enc_total}/{total} encontrados", "ok")
    return resultados

# PATCH para processor.py
# ============================================================
# INSTRUCAO: adicione este codigo NO FINAL do processor.py,
# antes da ultima linha (se houver alguma fora das funcoes).
# ============================================================



# ---------------------------------------------------------------------------
# Busca Individual — pesquisa por nome sem planilha
# ---------------------------------------------------------------------------

def buscar_individual(
    nome: str,
    drive_raiz: str,
    pasta_destino: str,
    data: str = "",
    threshold_fuzzy: float = 80.0,
    callback=None,
    modo_extracao: str = MODO_AUTO,
    dpi_ocr: int = 150,
    max_workers: int = 4,
    filtrar_aso: bool = False,
) -> dict:
    """
    Busca um único paciente pelo nome nos PDFs do drive.

    Muito mais simples que processar_lista() — sem planilha, sem agrupamentos.
    Ideal para o operador que quer encontrar um documento específico rapidinho.

    Parâmetros:
        nome          : nome completo (ou parcial) do paciente
        drive_raiz    : pasta raiz do drive com os PDFs
        pasta_destino : onde salvar o PDF encontrado
        data          : data no formato DD/MM, DD/MM/AAAA ou vazio para varrer tudo
        threshold_fuzzy: sensibilidade da busca (60-100)
        callback      : função(progresso, etapa, detalhe, status) para UI

    Retorna dict com:
        encontrado    : bool
        nome          : str
        data          : str
        arquivo       : str  (caminho do PDF salvo)
        score_fuzzy   : float
        erro          : str  (mensagem se não encontrou)
        pdfs_buscados : int  (quantos PDFs foram verificados)
    """
    def _cb(prog, etapa, detalhe="", status="info"):
        if callback:
            callback(prog, etapa, detalhe, status)

    resultado = {
        "encontrado"   : False,
        "nome"         : nome.strip(),
        "data"         : data.strip(),
        "arquivo"      : "",
        "score_fuzzy"  : 0.0,
        "erro"         : "",
        "pdfs_buscados": 0,
    }

    nome = nome.strip()
    if not nome:
        resultado["erro"] = "Informe o nome do paciente."
        return resultado

    _cb(0.05, "Localizando pasta...", drive_raiz)

    # ── Resolve a pasta de busca ────────────────────────────────
    data_val = data.strip() if data else None
    modo_reg = classificar_modo_registro(data_val) if data_val else "varredura"

    if modo_reg == "varredura" or not data_val:
        pasta = Path(drive_raiz)
        _cb(0.10, "Modo varredura", "Sem data — buscando em todos os PDFs", "aviso")
        pdfs = coletar_pdfs_recursivo(pasta)
    else:
        try:
            pasta, descricao = resolver_pasta_pdfs(drive_raiz, data_val)
            _cb(0.10, "Pasta localizada", descricao, "ok")
            if not pasta.exists():
                resultado["erro"] = f"Pasta não encontrada: {pasta.name}"
                _cb(1.0, "Pasta não encontrada", str(pasta), "erro")
                return resultado
            pdfs = sorted(pasta.glob("*.pdf"))
        except Exception as e:
            resultado["erro"] = f"Erro ao montar caminho: {e}"
            return resultado

    n_pdfs = len(pdfs)
    resultado["pdfs_buscados"] = n_pdfs

    if n_pdfs == 0:
        resultado["erro"] = "Nenhum PDF encontrado na pasta."
        _cb(1.0, "Sem PDFs", "Verifique o caminho do drive e a data.", "erro")
        return resultado

    _cb(0.15, "Buscando...", f"{n_pdfs} PDF(s) para verificar")

    # ── Busca nos PDFs em paralelo ──────────────────────────────
    melhor_pdf    = None
    melhor_pags   = []
    melhor_score  = 0.0
    em_descartada = 0
    workers       = min(max_workers, n_pdfs)

    import threading
    parar = threading.Event()

    with ThreadPoolExecutor(max_workers=workers) as ex:
        futuros = {
            ex.submit(
                _worker_pdf,
                j, pdf,
                [nome],
                threshold_fuzzy,
                modo_extracao,
                dpi_ocr,
                n_pdfs,
                filtrar_aso,
            ): (j, pdf)
            for j, pdf in enumerate(pdfs)
        }

        for fut in as_completed(futuros):
            if parar.is_set():
                break
            try:
                j, pdf, res, logs = fut.result()
            except Exception:
                continue

            prog = 0.15 + 0.75 * ((j + 1) / n_pdfs)
            _cb(prog, f"Verificando PDF {j+1}/{n_pdfs}", pdf.name)

            pags, score, desc = res.get(nome, ([], 0.0, 0))
            if score > melhor_score:
                melhor_score  = score
                melhor_pdf    = pdf
                melhor_pags   = pags
                em_descartada = desc

            # Encontrou com boa pontuação — para cedo
            if melhor_score >= 98.0:
                parar.set()
                for f in futuros:
                    f.cancel()
                break

    # ── Salva o resultado ───────────────────────────────────────
    if melhor_pags:
        try:
            saida = extrair_e_salvar_paginas(
                melhor_pdf, melhor_pags,
                Path(pasta_destino), nome,
            )
            resultado.update({
                "encontrado" : True,
                "arquivo"    : str(saida),
                "score_fuzzy": melhor_score,
            })
            _cb(1.0, "Documento encontrado!",
                f"{nome}  —  score {melhor_score:.0f}%", "ok")
        except Exception as e:
            resultado["erro"] = f"Erro ao salvar PDF: {e}"
            _cb(1.0, "Erro ao salvar", str(e), "erro")
    else:
        if filtrar_aso and em_descartada > 0:
            resultado["erro"] = (
                f"Nome encontrado em {em_descartada} página(s), "
                "mas nenhuma é ASO. Desative o filtro ASO para extrair."
            )
            _cb(1.0, "Encontrado, mas não é ASO",
                f"{em_descartada} pág(s) descartada(s) pelo filtro", "aviso")
        elif melhor_score > 0:
            resultado["erro"] = (
                f"Nome encontrado com score {melhor_score:.0f}% "
                f"(mínimo necessário: {threshold_fuzzy:.0f}%). "
                "Tente reduzir a sensibilidade."
            )
            _cb(1.0, "Score insuficiente",
                f"Melhor match: {melhor_score:.0f}% em {n_pdfs} PDF(s)", "aviso")
        else:
            resultado["erro"] = (
                f"Nome não localizado em {n_pdfs} PDF(s). "
                "Verifique a grafia ou a data informada."
            )
            _cb(1.0, "Não encontrado",
                f"Buscado em {n_pdfs} PDF(s)", "aviso")

    return resultado